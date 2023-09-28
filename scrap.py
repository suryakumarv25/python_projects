# from bs4 import BeautifulSoup
# import requests
# from openpyxl import Workbook

# #Defining the workbook and column headers 
# wb = Workbook()
# ws = wb.active

# ws['A1'] = "Match Day"
# ws['B1'] = "Match Time"
# ws['C1'] = "Visitor Team"
# ws['D1'] = "Home Team"
# ws['E1'] = "Arena"

# #Iterating through URL

    
# url = f"https://www.basketball-reference.com/leagues/NBA_2023_games-october.html"
# response = requests.get(url)
# html = response.text
# #reading the contents of each URL
# soup = BeautifulSoup(html, "html.parser")
# Schedule = soup.find('tbody')
# for match in Schedule.find_all('tr'):
#     match_day = match.find('th').text
#     match_time = match.find('td').text
#     list1 = [match_day, match_time]
#     for td in match.find_all('td', class_='left'):
#         list1.append(td.text)
        
#     ws.append(list1)


# # Save the file
# wb.save("sample.xlsx")


##################################################################

from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook

#Defining the workbook and column headers 
wb = Workbook()
ws = wb.active

ws['A1'] = "ROLL NUMBER"
ws['B1'] = "Player"
ws['C1'] = "Matches"
ws['D1'] = "Inns"
ws['E1'] = "Runs"
ws['F1'] = "Avg"
ws['G1'] = "Sr	"
ws['H1'] = "4s"
ws['I1'] = "6s"


#Iterating through URL

    
url = f"https://www.cricbuzz.com/cricket-series/6913/australia-tour-of-india-2023/stats"
response = requests.get(url)
html = response.text
#reading the contents of each URL
soup = BeautifulSoup(html, "html.parser")
Schedule = soup.find('tbody')
for match in Schedule.find_all('tr'):
    Player=match.find('td',class_="cb-srs-stats-td text-left")
    list1=[]
    for td in match.find_all('td',class_="cb-srs-stats-td text-right"):
        list1.append(td.text)
    list1.insert(1,Player.text)    
    
        
    ws.append(list1)


# Save the file
wb.save("sample.xlsx")
#test_url = "https://www.cricbuzz.com/cricket-series/6913/australia-tour-of-india-2023/stats"